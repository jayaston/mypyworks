# -*- coding: utf-8 -*-
"""
Created on Fri Apr 12 11:55:22 2019

@author: XieJie
"""
import sys
import os
try:
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__),"StatLedger\module")))
except:
    sys.path.append(r'.\mypyworks\StatLedger\module')
import pandas as pd
import numpy as np
import tjfxdata as tjfx
import openpyxl as oxl
import shutil
import os
#import re    

baobiaoriqi = '20200827'

list1 = [['00','00718','d'],
         ['1001','00718','d'],
         ['1002','00718','d'],
         ['1003','00718','d'],
         ['1016','00718','d'],
         ['1004','00718','d'],
         ['1005','00718','d'],
         ['1007','00718','d']]
shuju_df = tjfx.TjfxData().getdata(baobiaoriqi,baobiaoriqi,list1)

shuju_df.info()

shuju_df.QUOTA_VALUE = pd.to_numeric(shuju_df.QUOTA_VALUE,errors='coerce').fillna(0)
shuju_df.QUOTA_VALUE=shuju_df.QUOTA_VALUE/10000

file_dir = "E:/每日快报营运周报/"
file_name = '%s%s%s' % ('广州自来水每日快报' , baobiaoriqi , '.xlsx')
file_abs = os.path.join(file_dir, file_name)
os.chdir(file_dir)
shutil.copyfile("广州自来水每日快报190416.xlsx",file_name)  
writer = pd.ExcelWriter(file_abs, engine='openpyxl')
# 加载指定的excel文件
writer.book = oxl.load_workbook(file_abs)
 # 得到指定sheet的最后一行数据，因为是在原excel里面添加内容
    # 所以添加的信息应该从当前sheet最后一行的后面开始
startrow = 0
#if startrow == None and '原始数据' in writer.book.sheetnames:
#    startrow = writer.book['原始数据'].max_row
 
# 是否需要重新创建一下该sheet  
idx = writer.book.sheetnames.index('Sheet2')        
writer.book.remove(writer.book.worksheets[idx])        
writer.book.create_sheet('Sheet2', idx) 
# copy existing sheets
writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
 
#df.apply(axis = 1)  #axis = 1 指定逐行添加，如果axis = 0，就是逐列添加信息
shuju_df.to_excel(writer,'Sheet2', startrow=startrow, index=False, header=False)
 
writer.save()
writer.close()


#df2 = shuju_df.copy()
#with pd.ExcelWriter('E:/每日快报营运周报/广州自来水每日快报模板.xlsx') as writer: 
#    shuju_df.to_excel(writer, sheet_name='Sheet_name_1')
#    df2.to_excel(writer, sheet_name='Sheet_name_2')
#

#import xlrd
#import xlwt
#from xlutils.copy import copy
#workbook = xlrd.open_workbook(r'E:/每日快报营运周报/广州自来水每日快报模板.xls',formatting_info=True) 
#workbookc = copy(workbook)
#col = 0
#for item in shuju_df.columns.values.tolist():
#    columndata = shuju_df[item].tolist()
#    for i in range(len(columndata)):
#        workbookc.get_sheet(1).write(i,col,str(columndata[i]))
#    col += 1 
#workbookc.save(r'E:/每日快报营运周报/广州自来水每日快报190417.xls')
